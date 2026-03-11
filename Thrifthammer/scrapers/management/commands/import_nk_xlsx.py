"""
Management command: import_nk_xlsx

Imports Noble Knight prices from an Octoparse-generated XLSX file into
CurrentPrice records, matching spreadsheet rows to our products by name.

Noble Knight XLSX format (columns):
    Current (3-column) format from Octoparse:
        Title    - NK product name (used for name matching)
        URL      - Direct Noble Knight product URL
        Price2   - USD price (e.g. 58.95)

    Legacy (5-column) format also supported:
        SKU / Name   - NK product name
        Game         - Game system (e.g. "Warhammer 40,000")
        Price        - USD price
        Item Number  - NK internal product ID
        URL          - Direct Noble Knight product URL

    The format is auto-detected from the header column count.

Matching strategy:
    For each active product in our DB, the command searches the spreadsheet
    for the best name match. Both sides are normalised — faction prefixes
    stripped, punctuation removed, edition years ignored — then scored by
    F1-like word overlap. A minimum score (--min-score, default 0.5) is
    required before a match is accepted.

Usage:
    # Preview matches without saving:
    python manage.py import_nk_xlsx --file "path/to/Adepta Sororitas SKUs.xlsx" --dry-run

    # Import matching rows:
    python manage.py import_nk_xlsx --file "path/to/Adepta Sororitas SKUs.xlsx"

    # Process an entire directory of XLSX files:
    python manage.py import_nk_xlsx --dir "path/to/SKUs/"

    # Raise/lower the match confidence threshold:
    python manage.py import_nk_xlsx --file "..." --min-score 0.6

Notes:
    - Prices are stored as-is from the spreadsheet (USD). Noble Knight is a
      US retailer so prices will display in USD on the site.
    - Products with no match above --min-score are silently skipped.
    - Safe to re-run — uses update_or_create.
    - Each product gets exactly one NK CurrentPrice row. When multiple
      editions of the same product match (e.g. "2016" and "2021" boxes),
      the cheapest valid price is stored so ThriftHammer always shows
      the best available deal.
"""

import os
import re
import glob
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError

from prices.models import CurrentPrice
from products.models import Product, Retailer


# ── Normalisation constants ────────────────────────────────────────────────────

# Words stripped from both product names and NK titles before matching.
# Includes edition years and non-product words that inflate similarity scores.
_SKIP_WORDS = frozenset({
    'edition', 'new', 'box', 'set', 'the', 'and', 'of', 'at',
    '2014', '2015', '2016', '2017',  # older NK edition years
    '2018', '2019', '2020', '2021', '2022', '2023', '2024', '2025',
    'index', 'codex', 'datacards', 'datasheets',  # rulebooks — never match to miniatures
    # Generic unit-type words that appear across many different products;
    # stripping them prevents false cross-product matches.
    'squad', 'warriors', 'warrior', 'guard', 'veteran',
})

# Faction/game prefix phrases stripped before matching so that "adepta sororitas"
# in an NK title doesn't inflate the score for every Sororitas product equally.
_FACTION_PREFIXES = [
    # Warhammer 40,000 factions
    'adepta sororitas', 'adeptus astartes', 'adeptus custodes',
    'adeptus mechanicus', 'adeptus titanicus', 'aeldari', 'astra militarum',
    'black templars', 'blood angels', 'chaos daemons', 'chaos space marines',
    'craftworlds', 'dark angels', 'death guard', 'deathwatch', 'drukhari',
    'genestealer cults', 'grey knights', 'horus heresy', 'imperial guard',
    'imperial knights', 'kill team', 'leagues of votann', 'necromunda',
    'necrons', 'necron',       # both plural and singular forms
    'orks', 'ork',             # both plural and singular forms
    'skitarii',                # strips from both "Skitarii Rangers" and "Datacards - Skitarii"
    'space marines', 'space marine',  # both plural and singular
    'space wolves', 'thousand sons', "t'au empire", 'tau empire',
    "t'au", 'tau',             # bare prefix used in some product names
    'tyranids', 'tyranid',     # both plural and singular
    'ultramarines', 'warhammer 40000', 'warhammer 40k', 'world eaters',
    # Age of Sigmar factions (included so AoS products don't get common words counted)
    'age of sigmar', 'blades of khorne', 'cities of sigmar',
    'daughters of khaine', 'disciples of tzeentch', 'flesh-eater courts',
    'gloomspite gitz', 'lumineth realm-lords', 'maggotkin of nurgle',
    'nighthaunt', 'orruk warclans', 'ossiarch bonereapers',
    'skaven', 'slaves to darkness', 'stormcast eternals',
    # Other specialist games
    'necromunda', 'warcry',
]

NK_RETAILER_SLUG = 'noble-knight-games'


class Command(BaseCommand):
    """
    Import Noble Knight prices from an Octoparse XLSX into CurrentPrice records.

    Matches spreadsheet rows to our products by normalised name similarity.
    Safe to re-run (idempotent).
    """

    help = 'Import Noble Knight prices from an Octoparse XLSX file.'

    def add_arguments(self, parser):
        """Register command-line arguments."""
        group = parser.add_mutually_exclusive_group(required=True)
        group.add_argument(
            '--file',
            type=str,
            metavar='PATH',
            help='Path to a single XLSX file to import.',
        )
        group.add_argument(
            '--dir',
            type=str,
            metavar='DIR',
            help='Directory of XLSX files to import (processes all *.xlsx).',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            default=False,
            help='Preview matches without saving to the database.',
        )
        parser.add_argument(
            '--min-score',
            type=float,
            default=0.5,
            metavar='SCORE',
            help='Minimum match score to accept (0.0–1.0, default: 0.5).',
        )
        parser.add_argument(
            '--name-filter',
            type=str,
            default=None,
            metavar='TEXT',
            help=(
                'Only match DB products whose name contains TEXT '
                '(case-insensitive). Auto-detected from the filename if omitted '
                '— e.g. "Adepta Sororitas SKUs.xlsx" filters to "Adepta Sororitas".'
            ),
        )

    # ── Entry point ────────────────────────────────────────────────────────────

    def handle(self, *args, **options):
        """Load XLSX file(s) and import matching prices."""
        dry_run = options['dry_run']
        min_score = options['min_score']

        # Resolve file list
        if options['file']:
            xlsx_files = [options['file']]
        else:
            pattern = os.path.join(options['dir'], '*.xlsx')
            xlsx_files = sorted(glob.glob(pattern))
            if not xlsx_files:
                raise CommandError(f'No .xlsx files found in: {options["dir"]}')

        # Validate retailer exists
        try:
            nk_retailer = Retailer.objects.get(slug=NK_RETAILER_SLUG)
        except Retailer.DoesNotExist:
            raise CommandError(
                f'Retailer "{NK_RETAILER_SLUG}" not found in DB. '
                'Run populate_products first.'
            )

        # Load all active products once (filtered per file below)
        all_products = list(Product.objects.filter(is_active=True))
        global_name_filter = options.get('name_filter')

        self.stdout.write(f'\nNoble Knight XLSX Import')
        self.stdout.write('=' * 60)
        self.stdout.write(f'  Files      : {len(xlsx_files)}')
        self.stdout.write(f'  Products   : {len(all_products)} active')
        self.stdout.write(f'  Min score  : {min_score}')
        self.stdout.write(f'  Dry run    : {dry_run}')
        self.stdout.write('=' * 60 + '\n')

        total_imported = 0
        total_skipped = 0

        for xlsx_path in xlsx_files:
            imported, skipped = self._process_file(
                xlsx_path, all_products, nk_retailer, min_score, dry_run,
                name_filter=global_name_filter,
            )
            total_imported += imported
            total_skipped += skipped

        # Summary
        self.stdout.write('\n' + '=' * 60)
        self.stdout.write(self.style.SUCCESS('Summary'))
        self.stdout.write('=' * 60)
        self.stdout.write(self.style.SUCCESS(f'  Imported : {total_imported}'))
        self.stdout.write(f'  Skipped  : {total_skipped} (no match above threshold)')
        if dry_run:
            self.stdout.write(
                self.style.WARNING('\n  DRY RUN — no changes saved to database.')
            )
        self.stdout.write('=' * 60 + '\n')

    # ── File processing ────────────────────────────────────────────────────────

    def _process_file(self, xlsx_path, all_products, nk_retailer, min_score, dry_run,
                      name_filter=None):
        """
        Process a single XLSX file.

        Products are filtered to those whose name contains name_filter (or the
        faction auto-detected from the filename) so that e.g. a Sororitas file
        only matches Sororitas products — not every other faction's Combat Patrol.

        Returns (imported_count, skipped_count).
        """
        try:
            import openpyxl
        except ImportError:
            raise CommandError(
                'openpyxl is required: pip install openpyxl'
            )

        if not os.path.isfile(xlsx_path):
            self.stderr.write(self.style.ERROR(f'File not found: {xlsx_path}'))
            return 0, 0

        # Auto-detect faction from filename if no explicit filter given.
        # "Adepta Sororitas SKUs.xlsx" → filter products by "Adepta Sororitas".
        if name_filter is None:
            basename = os.path.basename(xlsx_path)
            name_filter = re.sub(r'\s*SKUs?\.xlsx$', '', basename, flags=re.IGNORECASE).strip()

        # Filter to relevant faction products only — prevents "Combat Patrol"
        # from matching every faction when processing a single-faction file.
        faction_keyword = name_filter.lower()
        products = [
            p for p in all_products
            if faction_keyword in p.name.lower()
        ]

        self.stdout.write(
            f'--- {os.path.basename(xlsx_path)} '
            f'(filter: "{name_filter}", {len(products)} products) ---'
        )

        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active

        # Detect column format from the header row.
        #   Old format (5 cols): (SKU/Name, Game, Price, Item Number, URL)
        #   New format (3 cols): (Title, URL, Price2)
        # Price is column index 2 in both; URL position differs.
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        ncols = sum(1 for c in header if c is not None)
        new_format = (ncols == 3)  # True → (title, url, price); False → old 5-col

        nk_rows = [
            row for row in ws.iter_rows(min_row=2, values_only=True)
            if row[0]  # skip blank rows
        ]
        wb.close()

        if not nk_rows:
            self.stdout.write(self.style.WARNING('  No data rows found — skipping.'))
            return 0, 0

        imported = 0
        matched_pks = set()

        for product in products:
            candidates = self._all_matches(product.name, nk_rows, min_score)

            if not candidates:
                continue  # no match in this file for this product

            # Default: highest-scoring match (used when no candidate has a price).
            chosen_score, chosen_row = candidates[0]
            chosen_price = None

            # Among near-top-scoring candidates only, pick the cheapest valid price.
            # SCORE_TOLERANCE ensures we only compare true edition duplicates
            # (same score) — not different products that share one keyword.
            # e.g. "Assault Intercessors" (1.0) beats "Heavy Intercessors" (0.5)
            # even if Heavy Intercessors is cheaper.
            # Older editions with the same score are fairly compared (2016 vs 2021).
            SCORE_TOLERANCE = 0.10
            best_score = candidates[0][0]
            for score, row in candidates:
                if best_score - score > SCORE_TOLERANCE:
                    break  # remaining candidates diverge too far; stop
                p = self._parse_price(row[2])
                if p is not None and (chosen_price is None or p < chosen_price):
                    chosen_price = p
                    chosen_score, chosen_row = score, row

            edition_note = (
                f' ({len(candidates)} editions found, cheapest chosen)'
                if len(candidates) > 1 else ''
            )

            # Unpack the chosen row according to the detected column format.
            if new_format:
                # New 3-column format: (Title, URL, Price2)
                nk_name, url_raw, price_raw = chosen_row
            else:
                # Old 5-column format: (Name, Game, Price, Item Number, URL)
                nk_name, _game, price_raw, _item_num, url_raw = chosen_row

            price = chosen_price if chosen_price is not None else self._parse_price(price_raw)
            url = str(url_raw).strip() if url_raw else ''

            self.stdout.write(
                f'  [{chosen_score:.2f}] {product.name}{edition_note}\n'
                f'         -> {nk_name}  ${price_raw}  {url[:70]}'
            )

            if not dry_run:
                CurrentPrice.objects.update_or_create(
                    product=product,
                    retailer=nk_retailer,
                    defaults={
                        'price': price,
                        'url': url,
                        'listing_title': str(nk_name).strip() if nk_name else '',
                        'in_stock': True,
                        'not_available': False,
                    },
                )

            matched_pks.add(product.pk)
            imported += 1

        # Mark every in-scope product that had no spreadsheet match as "not
        # available" at Noble Knight. This ensures the website shows "Not
        # Available" (no link, no price) rather than stale placeholder data.
        unmatched = [p for p in products if p.pk not in matched_pks]
        not_available_count = len(unmatched)

        if not dry_run and unmatched:
            for product in unmatched:
                CurrentPrice.objects.update_or_create(
                    product=product,
                    retailer=nk_retailer,
                    defaults={
                        'price': None,
                        'url': '',
                        'listing_title': '',
                        'in_stock': False,
                        'not_available': True,
                    },
                )

        skipped = len(products) - imported

        self.stdout.write(
            f'  => {imported} matched, '
            f'{not_available_count} marked not available'
            + (' (dry run)' if dry_run else '')
            + '\n'
        )
        return imported, skipped

    # ── Name matching ──────────────────────────────────────────────────────────

    @staticmethod
    def _get_faction(text):
        """
        Return the first faction prefix found in text (lowercase), or None.

        Prefixes are checked longest-first so 'chaos space marines' is
        returned before 'space marines' for a matching text.
        """
        lower = text.lower()
        for prefix in sorted(_FACTION_PREFIXES, key=len, reverse=True):
            if prefix in lower:
                return prefix
        return None

    @staticmethod
    def _all_matches(product_name, nk_rows, min_score):
        """
        Return all NK rows whose name similarity meets min_score.

        NK often lists multiple editions of the same product (e.g. "Acolyte
        Hybrids 2016" and "Acolyte Hybrids 2021") at different prices. This
        method collects every candidate so the caller can pick the cheapest.

        A faction-conflict penalty (-0.6) is applied when our product belongs
        to faction A but the NK row explicitly names a different faction B.
        This prevents "Combat Patrol - Blood Angels" from being matched to
        "Dark Angels Combat Patrol" — the penalty drops the score below the
        default min_score (0.5) so the wrong-faction row is excluded entirely.

        Returns a list of (score, row) tuples sorted by score descending,
        or an empty list if no row meets min_score.
        """
        p_words = Command._keywords(product_name)
        if not p_words:
            return []

        # Detect which faction (if any) our DB product belongs to.
        our_faction = Command._get_faction(product_name)

        matches = []
        for row in nk_rows:
            nk_title = str(row[0])
            nk_words = Command._keywords(nk_title)
            if not nk_words:
                continue
            shared = len(p_words & nk_words)
            score = 2 * shared / (len(p_words) + len(nk_words))

            # Faction-conflict penalty: if our product has faction A and the
            # NK row explicitly names a DIFFERENT faction B, subtract 0.6 so
            # the wrong-faction row falls below the min_score threshold and
            # is excluded. E.g. "Combat Patrol - Blood Angels" (NK) won't
            # match "Dark Angels Combat Patrol" (DB).
            if our_faction is not None:
                nk_faction = Command._get_faction(nk_title)
                if nk_faction is not None and nk_faction != our_faction:
                    score -= 0.6

            if score >= min_score:
                matches.append((score, row))

        return sorted(matches, key=lambda x: x[0], reverse=True)

    @staticmethod
    def _keywords(name):
        """
        Return the normalised keyword set for a product name.

        Strips faction prefixes, edition years, punctuation, and stop words,
        leaving only the meaningful distinguishing words for comparison.
        """
        text = name.lower()
        # Remove parenthetical editions: "(2021 Edition)", "(New)"
        text = re.sub(r'\(.*?\)', '', text)
        # Strip faction/game prefixes (longest first to avoid partial matches)
        for prefix in sorted(_FACTION_PREFIXES, key=len, reverse=True):
            text = text.replace(prefix, ' ')
        # Strip remaining punctuation (keep hyphens temporarily)
        text = re.sub(r'[^\w\s-]', ' ', text)
        text = re.sub(r'\s*-\s*', ' ', text)
        text = re.sub(r'\s+', ' ', text).strip()
        # Split and filter
        return {
            w for w in text.split()
            if len(w) > 2 and w not in _SKIP_WORDS
        }

    # ── Price parsing ──────────────────────────────────────────────────────────

    @staticmethod
    def _parse_price(raw):
        """
        Parse a price value from the spreadsheet cell.

        Returns Decimal or None.
        """
        if raw is None:
            return None
        try:
            cleaned = str(raw).replace('$', '').replace('£', '').replace(',', '').strip()
            price = Decimal(cleaned)
            if 0 < price < 2000:
                return price
        except (InvalidOperation, ValueError):
            pass
        return None
